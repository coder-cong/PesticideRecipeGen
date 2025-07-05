import asyncio
import json
from fastapi import FastAPI,Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from regex import P
from sympy import im
import torch
from transformers import AutoTokenizer, AutoModelForCausalLM, TextIteratorStreamer
from threading import Thread
import os
import httpx
import io
import openpyxl
from markdown import markdown
from bs4 import BeautifulSoup
import re
from typing import List, Dict
from urllib.parse import quote
import time
from openai import OpenAI

import os
os.environ['http_proxy'] = "http://202.199.13.107:10809"
os.environ['https_proxy'] = "http://202.199.13.107:10809"

device = "cuda" if torch.cuda.is_available() else "cpu"
print(f"Using device: {device}")

model_path = "/home/iiap/Simple_RLHF/model/actor_final"
#tokenizer = AutoTokenizer.from_pretrained(model_path)
#model = AutoModelForCausalLM.from_pretrained(model_path).to(device)
OPENAI_API_KEY = "2c7f3209-ac37-43aa-b0a4-3e1f3650188d"

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class Message(BaseModel):
    role: str
    content: str

class ChatCompletionRequest(BaseModel):
    model: str
    messages: List[Message]
    max_tokens: Optional[int] = 4096
    temperature: Optional[float] = 0.6

async def generate_response(prompt, max_tokens, temperature):
    inputs = tokenizer(prompt, return_tensors="pt").to(device)
    input_length = inputs.input_ids.shape[1]
    streamer = TextIteratorStreamer(tokenizer, skip_prompt=True, skip_special_tokens=True)
    
    generation_kwargs = dict(
        inputs,
        streamer=streamer,
        max_new_tokens=max_tokens,
        temperature=temperature,
    )

    thread = Thread(target=model.generate, kwargs=generation_kwargs)
    thread.start()

    for new_text in streamer:
        yield f"data: {json.dumps({'id': 'chatcmpl-123', 'object': 'chat.completion.chunk', 'model': 'your-model-name', 'created': 1726114622, 'choices': [{'index': 0, 'delta': {'content': new_text}, 'finish_reason': None}]})}\n\n"

    yield "data: [DONE]\n\n"


async def proxy_openai(request: ChatCompletionRequest):
    headers = {
        "Authorization": f"Bearer {OPENAI_API_KEY}",
        "Content-Type": "application/json",
    }
    
    url = "https://ark.cn-beijing.volces.com/api/v3/chat/completions"
    request_data = request.dict()
    request_data['stream'] = True  # 确保启用流式响应
    async with httpx.AsyncClient(timeout=None) as client:
        async with client.stream("POST", url, headers=headers, json=request_data) as response:
            async for line in response.aiter_lines():
                # 直接返回每一行
                yield f"{line}\n"

                
def extract_tables(text: str) -> List[Dict[str, List[List[str]]]]:
    tables = []
    table_pattern = r'\|.*\|'
    current_table = None
    current_title = None

    for line in text.split('\n'):
        if line.startswith('## '):
            current_title = line.strip('# ')
        elif re.match(table_pattern, line):
            if current_table is None:
                current_table = {'title': current_title, 'data': []}
            cells = [cell.strip() for cell in line.strip('|').split('|')]
            current_table['data'].append(cells)
        elif current_table is not None:
            tables.append(current_table)
            current_table = None

    if current_table is not None:
        tables.append(current_table)

    return tables

def extract_steps(text: str) -> List[str]:
    steps = []
    step_pattern = r'^\d+\.\s(.+)$'
    for line in text.split('\n'):
        match = re.match(step_pattern, line)
        if match:
            steps.append(match.group(1))
    return steps

@app.post("/v1/chat/completions")
async def create_chat_completion(request: ChatCompletionRequest):
    
    
    with open("promote.txt","r",encoding="utf-8") as file:
        content = file.read()
    message_data = {"role": "system", "content": content}
    
    system_promote= Message.model_validate(message_data)
    
    
    request.messages.insert(0,system_promote)
    
    print(request)

    request.temperature=0.6


    if request.model == 'llama3':
        # 使用您自己的模型
        prompt = "".join([f"{msg.role}: {msg.content}\n" for msg in request.messages])
        prompt += "assistant:"
        return StreamingResponse(generate_response(prompt, 2048, request.temperature), media_type="text/event-stream")
    
    else:
        # 将请求代理给OpenAI的API
        request.model="deepseek-v3-250324"
        return StreamingResponse(proxy_openai(request), media_type="text/event-stream")
        
    

@app.post('/export')
async def export_to_excel(data: dict = Body(...)):
    text = data.get('data', '')
    
    print(text)

    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认创建的sheet

    # 提取并处理表格
    tables = extract_tables(text)
    for table in tables:
        ws = wb.create_sheet(title=table['title'])
        for row in table['data']:
            ws.append(row)
    
    
    print("提取后table",tables)
    # 提取并处理步骤
    steps = extract_steps(text)
    if steps:
        ws = wb.create_sheet(title="制备步骤")
        for i, step in enumerate(steps, 1):
            ws.append([f"{i}. {step}"])
    print("提取后steps",steps)
    # 将Excel文件保存到内存中
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 使用 UTF-8 编码文件名
    filename = "配方.xlsx"
    encoded_filename = quote(filename)

    # 返回Excel文件
    headers = {
        'Content-Disposition': f'attachment; filename="{encoded_filename}"',
        'Access-Control-Expose-Headers': 'Content-Disposition'
    }
    
    

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8421)           

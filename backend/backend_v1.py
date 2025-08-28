import asyncio
import json
from fastapi import FastAPI, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from regex import P
from sympy import im
import torch
from transformers import AutoTokenizer, BitsAndBytesConfig, AutoModelForCausalLM, TextIteratorStreamer
from threading import Thread
import os
import httpx
import io
import openpyxl
from bs4 import BeautifulSoup
import re
from typing import List, Dict
from urllib.parse import quote
import time
from openai import OpenAI
from peft import PeftModel

import os
os.environ['http_proxy'] = "http://202.199.13.107:10809"
os.environ['https_proxy'] = "http://202.199.13.107:10809"

device = "cuda" if torch.cuda.is_available() else "cpu"
print(f"Using device: {device}")

# qwen路径配置
base_model_path = "/data/models/qwen2.5-72B"
#lora_adapter_path = "/root/projs/LLaMA-Factory/src/saves/Qwen2.5-72B-Instruct/lora/train_2025-07-05-14-48-49"

model_path = "/home/iiap/Simple_RLHF/model/actor_final"
# tokenizer = AutoTokenizer.from_pretrained(model_path)
# model = AutoModelForCausalLM.from_pretrained(model_path).to(device)
OPENAI_API_KEY = "2c7f3209-ac37-43aa-b0a4-3e1f3650188d"

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class Message(BaseModel):
    role: str
    content: str


class ChatCompletionRequest(BaseModel):
    model: str
    messages: List[Message]
    max_tokens: Optional[int] = 4096
    temperature: Optional[float] = 0.6


def load_model():
    bnb_config = BitsAndBytesConfig(
        load_in_4bit=True,
        bnb_4bit_quant_type="nf4",  # 使用NF4类型进行量化
        bnb_4bit_compute_dtype=torch.bfloat16,  # 在计算时，权重会反量化为bfloat16，以保持精度
        bnb_4bit_use_double_quant=True,  # 使用双重量化
    )

    tokenizer = AutoTokenizer.from_pretrained(base_model_path)

    print("正在加载基础模型到多张GPU上...")
    base_model = AutoModelForCausalLM.from_pretrained(
        base_model_path,
        torch_dtype=torch.bfloat16,  # 使用bfloat16以节省显存并提高性能 (如果你的GPU支持)
        quantization_config=bnb_config,  # 如果要使用4-bit量化，取消这一行的注释

        # -------- 关键参数 --------
        device_map="auto",
        # --------------------------

        # 如果遇到"out of memory"错误，可以尝试这个参数来减少峰值内存使用
        # low_cpu_mem_usage=True
    )

    print("基础模型加载完成。")

    # 注意：即使模型分布在多卡，输入数据通常也需要放到第一个设备上
    # `device_map="auto"`通常会将模型的输入层(word_embeddings)放在`cuda:0`上
    return (base_model, tokenizer)


async def qwen_generate_response(prompt, max_tokens, temperature):

    # --- 6. 进行推理 ---
    inputs = tokenizer.apply_chat_template(
        conversation=[{"role": "user", "content": prompt}], return_tensors="pt", add_generation_prompt=True)
    print(inputs)

    streamer = TextIteratorStreamer(
        tokenizer, skip_prompt=True, skip_special_tokens=True)

    generation_kwargs = dict(
        inputs=inputs,
        streamer=streamer,
        max_new_tokens=max_tokens,
        temperature=temperature
    )

    thread = Thread(target=model.generate, kwargs=generation_kwargs)
    thread.start()

    for new_text in streamer:
        yield f"data: {json.dumps({'id': 'chatcmpl-123', 'object': 'chat.completion.chunk', 'model': 'your-model-name', 'created': 1726114622, 'choices': [{'index': 0, 'delta': {'content': new_text}, 'finish_reason': None}]})}\n\n"

    yield "data: [None]"


async def generate_response(prompt, max_tokens, temperature):
    inputs = tokenizer(prompt, return_tensors="pt").to(device)
    input_length = inputs.input_ids.shape[1]
    streamer = TextIteratorStreamer(
        tokenizer, skip_prompt=True, skip_special_tokens=True)

    generation_kwargs = dict(
        inputs,
        streamer=streamer,
        max_new_tokens=max_tokens,
        temperature=temperature,
    )

    thread = Thread(target=model.generate, kwargs=generation_kwargs)
    thread.start()

    for new_text in streamer:
        yield f"data: {json.dumps({'id': 'chatcmpl-123', 'object': 'chat.completion.chunk', 'model': 'your-model-name', 'created': 1726114622, 'choices': [{'index': 0, 'delta': {'content': new_text}, 'finish_reason': None}]})}\n\n"

    yield "data: [DONE]\n\n"


async def proxy_openai(request: ChatCompletionRequest):
    headers = {
        "Authorization": f"Bearer {OPENAI_API_KEY}",
        "Content-Type": "application/json",
    }

    url = "https://ark.cn-beijing.volces.com/api/v3/chat/completions"
    request_data = request.dict()
    request_data['stream'] = True  # 确保启用流式响应
    async with httpx.AsyncClient(timeout=None) as client:
        async with client.stream("POST", url, headers=headers, json=request_data) as response:
            async for line in response.aiter_lines():
                # 直接返回每一行
                yield f"{line}\n"


def extract_tables(text: str) -> List[Dict[str, List[List[str]]]]:
    tables = []
    table_pattern = r'\|.*\|'
    current_table = None
    current_title = None

    for line in text.split('\n'):
        if line.startswith('## '):
            current_title = line.strip('# ')
        elif re.match(table_pattern, line):
            if current_table is None:
                current_table = {'title': current_title, 'data': []}
            cells = [cell.strip() for cell in line.strip('|').split('|')]
            current_table['data'].append(cells)
        elif current_table is not None:
            tables.append(current_table)
            current_table = None

    if current_table is not None:
        tables.append(current_table)

    return tables


def extract_steps(text: str) -> List[str]:
    steps = []
    step_pattern = r'^\d+\.\s(.+)$'
    for line in text.split('\n'):
        match = re.match(step_pattern, line)
        if match:
            steps.append(match.group(1))
    return steps


@app.post("/v1/chat/completions")
async def create_chat_completion(request: ChatCompletionRequest):

    with open("prompt.txt", "r", encoding="utf-8") as file:
        content = file.read()
    message_data = {"role": "system", "content": content}

    system_promote = Message.model_validate(message_data)

    request.messages.insert(0, system_promote)

    print(request)

    request.temperature = 0.6

    if request.model == 'qwen':
        # 使用您自己的模型
        prompt = "".join(
            [f"{msg.role}: {msg.content}\n" for msg in request.messages])
        prompt += "assistant:"
        return StreamingResponse(qwen_generate_response(prompt, 2048, request.temperature), media_type="text/event-stream")

    else:
        # 将请求代理给OpenAI的API
        request.model = "deepseek-v3-250324"
        return StreamingResponse(proxy_openai(request), media_type="text/event-stream")


@app.post('/export')
async def export_to_excel(data: dict = Body(...)):
    text = data.get('data', '')

    print(text)

    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认创建的sheet

    # 提取并处理表格
    tables = extract_tables(text)
    for table in tables:
        ws = wb.create_sheet(title=table['title'])
        for row in table['data']:
            ws.append(row)

    print("提取后table", tables)
    # 提取并处理步骤
    steps = extract_steps(text)
    if steps:
        ws = wb.create_sheet(title="制备步骤")
        for i, step in enumerate(steps, 1):
            ws.append([f"{i}. {step}"])
    print("提取后steps", steps)
    # 将Excel文件保存到内存中
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 使用 UTF-8 编码文件名
    filename = "配方.xlsx"
    encoded_filename = quote(filename)

    # 返回Excel文件
    headers = {
        'Content-Disposition': f'attachment; filename="{encoded_filename}"',
        'Access-Control-Expose-Headers': 'Content-Disposition'
    }

    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )


if __name__ == "__main__":
    import uvicorn
    model, tokenizer = load_model()
    uvicorn.run(app, host="0.0.0.0", port=8421)
